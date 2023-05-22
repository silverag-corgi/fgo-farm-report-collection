import glob
import os
from datetime import datetime
from enum import IntEnum, auto
from typing import Final, Optional, Union, cast

import openpyxl
import pandas as pd
import python_lib_for_me as pyl
from openpyxl.worksheet.worksheet import Worksheet
from styleframe import StyleFrame, Styler, utils

from fgo_farm_report_collection.util import const_util, pandas_util


class EnumOfProc(IntEnum):
    GENERATE_LIST = auto()
    GENERATE_USER_TOTAL_SUMMARY = auto()
    GENERATE_QUEST_TOTAL_SUMMARY = auto()
    GENERATE_INDIVIDUAL_SUMMARY = auto()


def do_logic_that_merge_list(
    use_debug_mode: bool,
    append_sheet: bool,
) -> None:
    """ロジック(周回報告一覧マージ)実行"""

    clg: Optional[pyl.CustomLogger] = None

    try:
        # ロガーの取得
        clg = pyl.CustomLogger(__name__, use_debug_mode=use_debug_mode)
        clg.log_inf(f"ロジック(周回報告一覧マージ)実行を開始します。")

        __do_logic_that_merge_gen_result(
            use_debug_mode,
            EnumOfProc.GENERATE_LIST,
            const_util.FARM_REPORT_LIST_FILE_PATH,
            const_util.FARM_REPORT_LIST_HEADER,
            append_sheet,
            const_util.FARM_REPORT_LIST_MERGE_RESULT_FILE_PATH,
            "周回報告一覧",
            {
                const_util.FARM_REPORT_LIST_HEADER[0]: 20,
                const_util.FARM_REPORT_LIST_HEADER[1]: 20,
                const_util.FARM_REPORT_LIST_HEADER[2]: 14,
                const_util.FARM_REPORT_LIST_HEADER[3]: 25,
                const_util.FARM_REPORT_LIST_HEADER[4]: 20,
                const_util.FARM_REPORT_LIST_HEADER[5]: 10,
                const_util.FARM_REPORT_LIST_HEADER[6]: 200,
            },
            "A4",
            ["A1:G1", "A2:G2"],
        )
    except Exception as e:
        raise (e)
    finally:
        if clg is not None:
            clg.log_inf(f"ロジック(周回報告一覧マージ)実行を終了します。")

    return None


def do_logic_that_merge_yearly_usr_tot_sum(
    use_debug_mode: bool,
    append_sheet: bool,
) -> None:
    """ロジック(周回報告年間ユーザ全体概要マージ)実行"""

    clg: Optional[pyl.CustomLogger] = None

    try:
        # ロガーの取得
        clg = pyl.CustomLogger(__name__, use_debug_mode=use_debug_mode)
        clg.log_inf(f"ロジック(周回報告年間ユーザ全体概要マージ)実行を開始します。")

        __do_logic_that_merge_gen_result(
            use_debug_mode,
            EnumOfProc.GENERATE_USER_TOTAL_SUMMARY,
            const_util.FARM_REPORT_YEARLY_USR_TOT_SUM_FILE_PATH,
            const_util.FARM_REPORT_USR_TOT_SUM_HEADER,
            append_sheet,
            const_util.FARM_REPORT_YEARLY_USR_TOT_SUM_MERGE_RESULT_FILE_PATH,
            "クエスト種別ごとの年間周回数(ユーザ編)",
            {
                const_util.FARM_REPORT_USR_TOT_SUM_HEADER[0]: 20,
                const_util.FARM_REPORT_USR_TOT_SUM_HEADER[1]: 20,
                const_util.FARM_REPORT_USR_TOT_SUM_HEADER[2]: 10,
                const_util.FARM_REPORT_USR_TOT_SUM_HEADER[3]: 10,
                const_util.FARM_REPORT_USR_TOT_SUM_HEADER[4]: 13,
                const_util.FARM_REPORT_USR_TOT_SUM_HEADER[5]: 13,
                const_util.FARM_REPORT_USR_TOT_SUM_HEADER[6]: 13,
                const_util.FARM_REPORT_USR_TOT_SUM_HEADER[7]: 13,
                const_util.FARM_REPORT_USR_TOT_SUM_HEADER[8]: 15,
            },
            "A4",
            ["A1:I1", "A2:I2"],
        )
    except Exception as e:
        raise (e)
    finally:
        if clg is not None:
            clg.log_inf(f"ロジック(周回報告年間ユーザ全体概要マージ)実行を終了します。")

    return None


def do_logic_that_merge_yearly_qst_tot_sum(
    use_debug_mode: bool,
    append_sheet: bool,
) -> None:
    """ロジック(周回報告年間クエスト全体概要マージ)実行"""

    clg: Optional[pyl.CustomLogger] = None

    try:
        # ロガーの取得
        clg = pyl.CustomLogger(__name__, use_debug_mode=use_debug_mode)
        clg.log_inf(f"ロジック(周回報告年間クエスト全体概要マージ)実行を開始します。")

        __do_logic_that_merge_gen_result(
            use_debug_mode,
            EnumOfProc.GENERATE_QUEST_TOTAL_SUMMARY,
            const_util.FARM_REPORT_YEARLY_QST_TOT_SUM_FILE_PATH,
            const_util.FARM_REPORT_QST_TOT_SUM_HEADER,
            append_sheet,
            const_util.FARM_REPORT_YEARLY_QST_TOT_SUM_MERGE_RESULT_FILE_PATH,
            "クエスト種別ごとの年間周回数(クエスト編)",
            {
                const_util.FARM_REPORT_QST_TOT_SUM_HEADER[0]: 20,
                const_util.FARM_REPORT_QST_TOT_SUM_HEADER[1]: 20,
                const_util.FARM_REPORT_QST_TOT_SUM_HEADER[2]: 10,
                const_util.FARM_REPORT_QST_TOT_SUM_HEADER[3]: 10,
                const_util.FARM_REPORT_QST_TOT_SUM_HEADER[4]: 13,
                const_util.FARM_REPORT_QST_TOT_SUM_HEADER[5]: 13,
                const_util.FARM_REPORT_QST_TOT_SUM_HEADER[6]: 13,
                const_util.FARM_REPORT_QST_TOT_SUM_HEADER[7]: 13,
                const_util.FARM_REPORT_QST_TOT_SUM_HEADER[8]: 15,
            },
            "A4",
            ["A1:I1", "A2:I2"],
        )
    except Exception as e:
        raise (e)
    finally:
        if clg is not None:
            clg.log_inf(f"ロジック(周回報告年間クエスト全体概要マージ)実行を終了します。")

    return None


def do_logic_that_merge_monthly_usr_tot_sum(
    use_debug_mode: bool,
    append_sheet: bool,
) -> None:
    """ロジック(周回報告月間ユーザ全体概要マージ)実行"""

    clg: Optional[pyl.CustomLogger] = None

    try:
        # ロガーの取得
        clg = pyl.CustomLogger(__name__, use_debug_mode=use_debug_mode)
        clg.log_inf(f"ロジック(周回報告月間ユーザ全体概要マージ)実行を開始します。")

        __do_logic_that_merge_gen_result(
            use_debug_mode,
            EnumOfProc.GENERATE_USER_TOTAL_SUMMARY,
            const_util.FARM_REPORT_MONTHLY_USR_TOT_SUM_FILE_PATH,
            const_util.FARM_REPORT_USR_TOT_SUM_HEADER,
            append_sheet,
            const_util.FARM_REPORT_MONTHLY_USR_TOT_SUM_MERGE_RESULT_FILE_PATH,
            "クエスト種別ごとの月間周回数(ユーザ編)",
            {
                const_util.FARM_REPORT_USR_TOT_SUM_HEADER[0]: 20,
                const_util.FARM_REPORT_USR_TOT_SUM_HEADER[1]: 20,
                const_util.FARM_REPORT_USR_TOT_SUM_HEADER[2]: 10,
                const_util.FARM_REPORT_USR_TOT_SUM_HEADER[3]: 10,
                const_util.FARM_REPORT_USR_TOT_SUM_HEADER[4]: 13,
                const_util.FARM_REPORT_USR_TOT_SUM_HEADER[5]: 13,
                const_util.FARM_REPORT_USR_TOT_SUM_HEADER[6]: 13,
                const_util.FARM_REPORT_USR_TOT_SUM_HEADER[7]: 13,
                const_util.FARM_REPORT_USR_TOT_SUM_HEADER[8]: 15,
            },
            "A4",
            ["A1:I1", "A2:I2"],
        )
    except Exception as e:
        raise (e)
    finally:
        if clg is not None:
            clg.log_inf(f"ロジック(周回報告月間ユーザ全体概要マージ)実行を終了します。")

    return None


def do_logic_that_merge_monthly_qst_tot_sum(
    use_debug_mode: bool,
    append_sheet: bool,
) -> None:
    """ロジック(周回報告月間クエスト全体概要マージ)実行"""

    clg: Optional[pyl.CustomLogger] = None

    try:
        # ロガーの取得
        clg = pyl.CustomLogger(__name__, use_debug_mode=use_debug_mode)
        clg.log_inf(f"ロジック(周回報告月間クエスト全体概要マージ)実行を開始します。")

        __do_logic_that_merge_gen_result(
            use_debug_mode,
            EnumOfProc.GENERATE_QUEST_TOTAL_SUMMARY,
            const_util.FARM_REPORT_MONTHLY_QST_TOT_SUM_FILE_PATH,
            const_util.FARM_REPORT_QST_TOT_SUM_HEADER,
            append_sheet,
            const_util.FARM_REPORT_MONTHLY_QST_TOT_SUM_MERGE_RESULT_FILE_PATH,
            "クエスト種別ごとの月間周回数(クエスト編)",
            {
                const_util.FARM_REPORT_QST_TOT_SUM_HEADER[0]: 20,
                const_util.FARM_REPORT_QST_TOT_SUM_HEADER[1]: 20,
                const_util.FARM_REPORT_QST_TOT_SUM_HEADER[2]: 10,
                const_util.FARM_REPORT_QST_TOT_SUM_HEADER[3]: 10,
                const_util.FARM_REPORT_QST_TOT_SUM_HEADER[4]: 13,
                const_util.FARM_REPORT_QST_TOT_SUM_HEADER[5]: 13,
                const_util.FARM_REPORT_QST_TOT_SUM_HEADER[6]: 13,
                const_util.FARM_REPORT_QST_TOT_SUM_HEADER[7]: 13,
                const_util.FARM_REPORT_QST_TOT_SUM_HEADER[8]: 15,
            },
            "A4",
            ["A1:I1", "A2:I2"],
        )
    except Exception as e:
        raise (e)
    finally:
        if clg is not None:
            clg.log_inf(f"ロジック(周回報告月間クエスト全体概要マージ)実行を終了します。")

    return None


def do_logic_that_merge_ind_sum(
    use_debug_mode: bool,
    append_sheet: bool,
) -> None:
    """ロジック(周回報告個人概要マージ)実行"""

    clg: Optional[pyl.CustomLogger] = None

    try:
        # ロガーの取得
        clg = pyl.CustomLogger(__name__, use_debug_mode=use_debug_mode)
        clg.log_inf(f"ロジック(周回報告個人概要マージ)実行を開始します。")

        __do_logic_that_merge_gen_result(
            use_debug_mode,
            EnumOfProc.GENERATE_INDIVIDUAL_SUMMARY,
            const_util.FARM_REPORT_IND_SUM_FILE_PATH,
            const_util.FARM_REPORT_IND_SUM_HEADER,
            append_sheet,
            const_util.FARM_REPORT_IND_SUM_MERGE_RESULT_FILE_PATH,
            "ユーザごとの周回数",
            {
                const_util.FARM_REPORT_IND_SUM_HEADER[0]: 10,
                const_util.FARM_REPORT_IND_SUM_HEADER[1]: 17,
                const_util.FARM_REPORT_IND_SUM_HEADER[2]: 17,
                const_util.FARM_REPORT_IND_SUM_HEADER[3]: 17,
            },
            "A4",
            ["A1:D1", "A2:D2"],
        )
    except Exception as e:
        raise (e)
    finally:
        if clg is not None:
            clg.log_inf(f"ロジック(周回報告個人概要マージ)実行を終了します。")

    return None


def __do_logic_that_merge_gen_result(
    use_debug_mode: bool,
    enum_of_proc: EnumOfProc,
    gen_result_file_path_format: str,
    gen_result_header: list[str],
    append_merge_result_sheet: bool,
    merge_result_file_path: str,
    merge_result_book_name: str,
    merge_result_column_widths: dict[str, Union[int, float]],
    merge_result_cell_to_fix_window_frame: str,
    merge_result_ranges_to_merge_cells: list[str],
) -> None:
    """ロジック(周回報告生成結果マージ(共通))実行"""

    clg: Optional[pyl.CustomLogger] = None

    try:
        # ロガーの取得
        clg = pyl.CustomLogger(__name__, use_debug_mode=use_debug_mode)

        # 周回報告生成結果ファイルパスの取得
        gen_result_file_dir: str = os.path.dirname(gen_result_file_path_format)
        gen_result_file_ext: str = os.path.splitext(gen_result_file_path_format)[1]
        gen_result_file_path_with_wildcard: str = gen_result_file_dir + "/*" + gen_result_file_ext
        gen_result_file_paths: list[str] = glob.glob(gen_result_file_path_with_wildcard)

        # 周回報告生成結果ファイルの件数が0件の場合
        if len(gen_result_file_paths) == 0:
            clg.log_wrn(
                f"周回報告生成結果ファイルの件数が0件です。(gen_result_file_path:{gen_result_file_path_with_wildcard})"
            )
        else:
            # Pandasによる周回報告マージ結果ファイルの生成
            __generate_merge_result_file_by_pandas(
                use_debug_mode,
                enum_of_proc,
                gen_result_file_paths,
                gen_result_header,
                append_merge_result_sheet,
                merge_result_file_path,
                merge_result_book_name,
                merge_result_column_widths,
                merge_result_cell_to_fix_window_frame,
            )

            # OpenPyXLによる周回報告マージ結果ファイルの編集
            __edit_merge_result_file_by_openpyxl(
                use_debug_mode, merge_result_file_path, merge_result_ranges_to_merge_cells
            )

            clg.log_inf(f"周回報告マージ結果ファイルパス：\n{merge_result_file_path}")
    except Exception as e:
        raise (e)

    return None


def __generate_merge_result_file_by_pandas(
    use_debug_mode: bool,
    enum_of_proc: EnumOfProc,
    gen_result_file_paths: list[str],
    gen_result_header: list[str],
    append_merge_result_sheet: bool,
    merge_result_file_path: str,
    merge_result_book_name: str,
    merge_result_column_widths: dict[str, Union[int, float]],
    merge_result_cell_to_fix_window_frame: str,
) -> None:
    """周回報告マージ結果ファイル生成(Pandas)"""

    clg: Optional[pyl.CustomLogger] = None

    try:
        # ロガーの取得
        clg = pyl.CustomLogger(__name__, use_debug_mode=use_debug_mode)

        excel_writer: Optional[pd.ExcelWriter] = None
        try:
            # Excelライターの生成
            if append_merge_result_sheet is True and os.path.isfile(merge_result_file_path) is True:
                excel_writer = StyleFrame.ExcelWriter(
                    merge_result_file_path,
                    mode="a",
                    if_sheet_exists="replace",  # overlay not working
                )
            else:
                excel_writer = StyleFrame.ExcelWriter(
                    merge_result_file_path,
                    mode="w",
                )

            # 周回報告マージ結果ファイルへの出力
            for gen_result_file_path in reversed(gen_result_file_paths):
                clg.log_inf(f"周回報告生成結果ファイルパス：\n{gen_result_file_path}")

                # 周回報告マージ結果シート名の生成
                merge_result_sheet_name: str = pyl.generate_file_name(gen_result_file_path)

                # 周回報告マージ結果ヘッダ部スタイルフレームの生成
                merge_result_header_part_sfs: list[
                    StyleFrame
                ] = __generate_merge_result_header_part_sfs(
                    use_debug_mode,
                    merge_result_book_name,
                    merge_result_sheet_name,
                    gen_result_header,
                )

                # 周回報告生成結果データフレームの取得
                gen_result_df: pd.DataFrame = pd.DataFrame()
                if enum_of_proc == EnumOfProc.GENERATE_LIST:
                    gen_result_df = pandas_util.read_farm_report_list_file(
                        use_debug_mode, gen_result_file_path, True, True
                    )
                elif enum_of_proc == EnumOfProc.GENERATE_USER_TOTAL_SUMMARY:
                    gen_result_df = pandas_util.read_farm_report_usr_tot_sum_file(
                        use_debug_mode, gen_result_file_path, True, True
                    )
                elif enum_of_proc == EnumOfProc.GENERATE_QUEST_TOTAL_SUMMARY:
                    gen_result_df = pandas_util.read_farm_report_qst_tot_sum_file(
                        use_debug_mode, gen_result_file_path, True, True
                    )
                elif enum_of_proc == EnumOfProc.GENERATE_INDIVIDUAL_SUMMARY:
                    gen_result_df = pandas_util.read_farm_report_ind_sum_file(
                        use_debug_mode, gen_result_file_path, True, True
                    )

                # 周回報告生成結果データフレームへの書式設定の適用
                # (周回報告マージ結果データ部スタイルフレームの生成)
                merge_result_data_part_sf: StyleFrame = __apply_formatting_to_gen_result(
                    use_debug_mode, gen_result_df, merge_result_column_widths
                )

                # 周回報告マージ結果スタイルフレームの保存
                pandas_util.save_farm_report_merge_result_sf(
                    use_debug_mode,
                    merge_result_header_part_sfs,
                    merge_result_data_part_sf,
                    excel_writer,
                    merge_result_sheet_name,
                    cell_to_fix_window_frame=merge_result_cell_to_fix_window_frame,
                )
        except Exception as e:
            clg.log_err(f"周回報告生成結果マージ結果ファイルへの出力に失敗しました。")
            raise (e)
        finally:
            if excel_writer is not None:
                excel_writer.close()
    except Exception as e:
        raise (e)

    return None


def __generate_merge_result_header_part_sfs(
    use_debug_mode: bool,
    merge_result_book_name: str,
    merge_result_sheet_name: str,
    gen_result_header: list[str],
) -> list[StyleFrame]:
    """周回報告マージ結果ヘッダ部スタイルフレーム生成"""

    merge_result_header_part_sfs: list[StyleFrame] = []

    # 周回報告マージ結果ヘッダ部01データフレームの生成
    merge_result_header_part_col_num: int = 0
    merge_result_header_part_01_df: pd.DataFrame = pd.DataFrame(
        {
            f"col_{merge_result_header_part_col_num}": [
                merge_result_book_name,
                merge_result_sheet_name,
            ]
        }
    )
    for _ in range(len(gen_result_header) - 1):
        merge_result_header_part_col_num = merge_result_header_part_col_num + 1
        merge_result_header_part_01_df[f"col_{merge_result_header_part_col_num}"] = ""

    # 周回報告マージ結果ヘッダ部01の書式設定の適用
    merge_result_header_part_01_sf: StyleFrame = __apply_formatting_to_merge_result_header_part(
        use_debug_mode, merge_result_header_part_01_df, is_update_datetime_col=False
    )

    # 更新日時の取得
    update_datetime: datetime = datetime.now()
    update_date: str = update_datetime.strftime("%Y-%m-%d")
    update_time: str = update_datetime.strftime("%H:%M:%S")

    # 周回報告マージ結果ヘッダ部02データフレームの生成
    merge_result_header_part_col_num = merge_result_header_part_col_num + 1
    merge_result_header_part_02_df: pd.DataFrame = pd.DataFrame(
        {f"col_{merge_result_header_part_col_num}": [update_date, update_time]}
    )

    # 周回報告マージ結果ヘッダ部02の書式設定の適用
    merge_result_header_part_02_sf: StyleFrame = __apply_formatting_to_merge_result_header_part(
        use_debug_mode, merge_result_header_part_02_df, is_update_datetime_col=True
    )

    # 周回報告マージ結果ヘッダ部スタイルフレームへの追加
    merge_result_header_part_sfs.append(merge_result_header_part_01_sf)
    merge_result_header_part_sfs.append(merge_result_header_part_02_sf)

    return merge_result_header_part_sfs


def __apply_formatting_to_merge_result_header_part(
    use_debug_mode: bool,
    merge_result_header_part_df: pd.DataFrame,
    merge_result_column_widths: dict[str, Union[int, float]] = {},
    is_update_datetime_col: bool = False,
) -> StyleFrame:
    """書式設定適用(周回報告マージ結果ヘッダ部)"""

    DATETIME_FORMAT: Final[str] = "YYYY-MM-DD HH:MM:SS"

    # デフォルトのスタイルの適用
    default_style: Styler = Styler(
        bg_color=None,
        bold=False,
        font=const_util.FONT_NAME,
        font_size=const_util.FONT_SIZE,
        number_format=utils.number_formats.general,
        horizontal_alignment=(
            utils.horizontal_alignments.general
            if is_update_datetime_col is False
            else utils.horizontal_alignments.right
        ),
        wrap_text=False,
        shrink_to_fit=False if is_update_datetime_col is False else True,
        date_time_format=DATETIME_FORMAT,
    )
    merge_result_header_part_sf: StyleFrame = StyleFrame(merge_result_header_part_df, default_style)

    # 列の幅の適用
    merge_result_header_part_sf.set_column_width_dict(merge_result_column_widths)

    # 行の高さの適用
    row_indexes: tuple = merge_result_header_part_sf.row_indexes
    merge_result_header_part_sf.set_row_height(row_indexes[: len(row_indexes) - 1], 20)

    return merge_result_header_part_sf


def __apply_formatting_to_gen_result(
    use_debug_mode: bool,
    gen_result_df: pd.DataFrame,
    merge_result_column_widths: dict[str, Union[int, float]] = {},
) -> StyleFrame:
    """書式設定適用(生成結果)"""

    DATETIME_FORMAT: Final[str] = "YYYY-MM-DD HH:MM:SS"

    # デフォルトのスタイルの適用
    default_style: Styler = Styler(
        bg_color=None,
        bold=False,
        font=const_util.FONT_NAME,
        font_size=const_util.FONT_SIZE,
        number_format=utils.number_formats.thousands_comma_sep,
        horizontal_alignment=utils.horizontal_alignments.general,
        wrap_text=False,
        shrink_to_fit=False,
        date_time_format=DATETIME_FORMAT,
    )
    merge_result_data_part_sf: StyleFrame = StyleFrame(gen_result_df, default_style)

    # ヘッダのスタイルの適用
    header_style: Styler = Styler(
        bg_color=utils.colors.grey,
        bold=True,
        font=const_util.FONT_NAME,
        font_size=const_util.FONT_SIZE,
        number_format=utils.number_formats.general,
        horizontal_alignment=utils.horizontal_alignments.center,
        wrap_text=True,
        shrink_to_fit=False,
        date_time_format=DATETIME_FORMAT,
    )
    merge_result_data_part_sf.apply_headers_style(header_style)

    # 列の幅の適用
    merge_result_data_part_sf.set_column_width_dict(merge_result_column_widths)

    # 行の高さの適用
    row_indexes: tuple = merge_result_data_part_sf.row_indexes
    if len(row_indexes) == 1:
        merge_result_data_part_sf.set_row_height(row_indexes[0], 30)
    else:
        merge_result_data_part_sf.set_row_height(row_indexes[0], 30)
        merge_result_data_part_sf.set_row_height(row_indexes[1:], 20)

    return merge_result_data_part_sf


def __edit_merge_result_file_by_openpyxl(
    use_debug_mode: bool,
    merge_result_file_path: str,
    merge_result_ranges_to_merge_cells: list[str],
) -> None:
    """周回報告マージ結果ファイル編集(OpenPyXL)"""

    clg: Optional[pyl.CustomLogger] = None

    try:
        # ロガーの取得
        clg = pyl.CustomLogger(__name__, use_debug_mode=use_debug_mode)

        merge_result_wb: Optional[openpyxl.Workbook] = None
        try:
            # 周回報告マージ結果ファイルのセルの結合
            merge_result_wb = openpyxl.load_workbook(merge_result_file_path)
            for sheet_name in merge_result_wb.sheetnames:
                merge_result_ws: Worksheet = cast(Worksheet, merge_result_wb[sheet_name])
                for range in merge_result_ranges_to_merge_cells:
                    merge_result_ws.merge_cells(range)
        except Exception as e:
            clg.log_err(f"周回報告生成結果マージ結果ファイルのセルの結合に失敗しました。")
            raise (e)
        finally:
            if merge_result_wb is not None:
                merge_result_wb.save(merge_result_file_path)
                merge_result_wb.close()
    except Exception as e:
        raise (e)

    return None
